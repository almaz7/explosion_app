from telebot import types
import datetime
import string_constants
from msg_check import check
import markups
import config
import visiting_fact
import users
import train_schedule
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils.cell import get_column_letter

bot = config.bot


def get_all_report(message: types.Message):
    bot.send_message(message.chat.id, "Введите дату, начиная с которой хотите получить отчет, в формате ДД-ММ-ГГГГ или напишите: Отмена")
    bot.register_next_step_handler(message, get_date_for_all_report)


def get_date_for_all_report(message: types.Message):
    status = check(str(message.text))
    if status == 1:
        bot.send_message(message.chat.id, string_constants.CANCEL_MSG,
                         reply_markup=markups.main_markup(message))
    else:
        if valid_date(message.text):
            bot.send_message(message.chat.id, "Превосходная дата!", reply_markup=markups.main_markup(message))
            date_str = message.text
            show_all_report(message, date_str)
        else:
            bot.send_message(message.chat.id, "Пожалуйста, введите корректную дату в формате ДД-ММ-ГГГГ или напишите: Отмена")
            bot.register_next_step_handler(message, get_date_for_all_report)


def show_all_report(message: types.Message, date_str: str):
    current_date = datetime.datetime.today()
    start_date = datetime.datetime.strptime(date_str, "%d-%m-%Y")
    date = start_date
    if date > current_date:
        bot.send_message(message.chat.id, "К сожалению, бот не может заглядывать в будущее. Скоро исправим)", reply_markup=markups.main_markup(message))
        return
    if date < (current_date - datetime.timedelta(days=365*2)):
        bot.send_message(message.chat.id, "Это было слишком давно. Бот уже не помнит события тех дней...",
                         reply_markup=markups.main_markup(message))
        return
    wb = Workbook()
    ws = wb.create_sheet("Посещения", 0)
    wb.create_sheet("Причины пропусков", 1)
    wb.remove(wb['Sheet'])

    make_all_sheet1(message, date_str, ws)
    ws = wb.worksheets[1]
    make_all_sheet2(message, date_str, ws)

    wb.save("report.xlsx")
    bot.send_message(message.chat.id, "Ваш отчет:",
                     reply_markup=markups.main_markup(message))
    wb.close()
    fd = open('report.xlsx', 'rb')
    bot.send_document(message.chat.id, fd)
    fd.close()


def make_all_sheet1(message: types.Message, date_str: str, ws: Workbook.worksheets):
    current_date = datetime.datetime.today()
    start_date = datetime.datetime.strptime(date_str, "%d-%m-%Y")
    date = start_date
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 25
    i = 2
    j = 1
    active_users = users.get_active_users_sorted_inBase()
    if not active_users:
        bot.send_message(message.chat.id, "В базе данных нет активных пользователей",
                         reply_markup=markups.main_markup(message))
        return
    ws.cell(row=1, column=3, value="Группа")
    for user in active_users:
        name = user[2] + " " + user[1]
        ws.cell(row=i, column=j, value=i - 1)
        ws.cell(row=i, column=j + 1, value=name)
        ws.cell(row=i, column=j + 2, value=user[8])
        if user[9] == 1:
            ws.cell(row=i, column=j + 1).fill = PatternFill('solid', fgColor="63C5DA")
            ws.cell(row=i, column=j + 2).fill = PatternFill('solid', fgColor="63C5DA")
        i += 1

    j = 4
    while date < current_date:
        rows = visiting_fact.get_visit_fact_row(date.strftime("%d-%m-%Y"))
        if not rows:
            date += datetime.timedelta(days=1)
            continue
        i = 1
        ws.column_dimensions[get_column_letter(j)].width = 11
        ws.cell(row=i, column=j, value=date.strftime("%d-%m-%Y")).alignment = Alignment(horizontal="center",
                                                                                        vertical="center")
        i = 2
        for user in active_users:
            for row in rows:
                if user[0] == row[0]:
                    if row[3] == 1:
                        answer = "-"
                        plan_row = visiting_fact.get_plan_answer_by_date_and_id(user[0], date.strftime("%d-%m-%Y"))
                        if plan_row and plan_row[0] == 2:
                            pf = PatternFill('solid', fgColor="F4CA16")
                        else:
                            pf = PatternFill('solid', fgColor="D7002F")
                    else:
                        answer = "+"
                        pf = PatternFill('solid', fgColor="50C878")
                    ws.cell(row=i, column=j, value=answer).alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(row=i, column=j).fill = pf
                    break
            i += 1
        j += 1
        date += datetime.timedelta(days=1)

    max_column = ws.max_column
    max_row = ws.max_row
    ws.column_dimensions[get_column_letter(max_column + 1)].width = 13
    ws.column_dimensions[get_column_letter(max_column + 2)].width = 13
    ws.cell(row=1, column=max_column + 1, value="Посещений").alignment = Alignment(horizontal="center",
                                                                                   vertical="center")
    ws.cell(row=1, column=max_column + 2, value="Пропусков").alignment = Alignment(horizontal="center",
                                                                                   vertical="center")
    for i in range(2, max_row + 1):
        count_here = 0
        count_not_here = 0
        for j in range(3, max_column + 1):
            if ws.cell(row=i, column=j).value == "+":
                count_here += 1
            elif ws.cell(row=i, column=j).value == "-":
                count_not_here += 1
        ws.cell(row=i, column=max_column + 1, value=count_here)
        ws.cell(row=i, column=max_column + 2, value=count_not_here)

    i = max_row + 2
    s1 = "Зелёный фон: был на занятии"
    s2 = "Жёлтый фон: отсутствие с причиной"
    s3 = "Красный фон: отсутствие без причины"
    s4 = "Голубой фон: в основном составе"
    ws.cell(row=i, column=2, value=s1).font = Font(color="50C878")
    ws.cell(row=i + 1, column=2, value=s2).font = Font(color="F4CA16")
    ws.cell(row=i + 2, column=2, value=s3).font = Font(color="D7002F")
    ws.cell(row=i + 3, column=2, value=s4).font = Font(color="63C5DA")


def make_all_sheet2(message: types.Message, date_str: str, ws: Workbook.worksheets):
    current_date = datetime.datetime.today()
    start_date = datetime.datetime.strptime(date_str, "%d-%m-%Y")
    date = start_date
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 25
    i = 2
    j = 1
    active_users = users.get_active_users_sorted_inBase()
    if not active_users:
        bot.send_message(message.chat.id, "В базе данных нет активных пользователей",
                         reply_markup=markups.main_markup(message))
        return
    for user in active_users:
        name = user[2] + " " + user[1]
        ws.cell(row=i, column=j, value=i - 1)
        ws.cell(row=i, column=j + 1, value=name)
        i += 1

    j = 3
    while date < current_date:
        fact_rows = visiting_fact.get_visit_fact_row(date.strftime("%d-%m-%Y"))
        if not fact_rows:
            date += datetime.timedelta(days=1)
            continue
        i = 1
        ws.column_dimensions[get_column_letter(j)].width = 15
        ws.cell(row=i, column=j, value=date.strftime("%d-%m-%Y")).alignment = Alignment(horizontal="center",
                                                                                        vertical="center")
        i = 2
        for user in active_users:
            for fact_row in fact_rows:
                if user[0] == fact_row[0]:
                    if fact_row[3] == 1: #отсутствовал
                        plan_row = visiting_fact.get_plan_answer_by_date_and_id(user[0], date.strftime("%d-%m-%Y"))
                        pf = PatternFill('solid', fgColor="F4CA16")
                        if plan_row and plan_row[0] == 2:
                            skip_reason = plan_row[1]
                        else:
                            skip_reason = "-"
                            pf = PatternFill('solid', fgColor="D7002F")
                        ws.cell(row=i, column=j, value=skip_reason).alignment = Alignment(horizontal="center", vertical="center")
                        ws.cell(row=i, column=j, value=skip_reason).fill = pf
                    break
            i += 1
        j += 1
        date += datetime.timedelta(days=1)


def get_my_report(message: types.Message):
    bot.send_message(message.chat.id, "Введите дату, начиная с которой хотите получить отчет, в формате ДД-ММ-ГГГГ или напишите: Отмена")
    bot.register_next_step_handler(message, get_date_for_my_report)


def get_date_for_my_report(message: types.Message):
    status = check(str(message.text))
    if status == 1:
        bot.send_message(message.chat.id, string_constants.CANCEL_MSG,
                         reply_markup=markups.main_markup(message))
    else:
        if valid_date(message.text):
            bot.send_message(message.chat.id, "Превосходная дата!", reply_markup=markups.main_markup(message))
            date_str = message.text
            show_my_report(message, date_str)
        else:
            bot.send_message(message.chat.id,
                             "Пожалуйста, введите корректную дату в формате ДД-ММ-ГГГГ или напишите: Отмена")
            bot.register_next_step_handler(message, get_date_for_my_report)


def show_my_report(message: types.Message, date_str: str):
    current_date = datetime.datetime.today()
    start_date = datetime.datetime.strptime(date_str, "%d-%m-%Y")
    date = start_date

    if date > current_date:
        bot.send_message(message.chat.id, "К сожалению, бот не может заглядывать в будущее. Скоро исправим)", reply_markup=markups.main_markup(message))
        return
    if date < (current_date - datetime.timedelta(days=365*2)):
        bot.send_message(message.chat.id, "Это было слишком давно. Бот уже не помнит события тех дней...",
                         reply_markup=markups.main_markup(message))
        return
    wb = Workbook()
    ws = wb.create_sheet("Посещения", 0)
    wb.create_sheet("Причины пропусков", 1)
    wb.remove(wb['Sheet'])

    make_my_sheet1(message, date_str, ws)
    ws = wb.worksheets[1]
    make_my_sheet2(message, date_str, ws)

    wb.save("my_report.xlsx")
    bot.send_message(message.chat.id, "Ваш персональный отчет:",
                     reply_markup=markups.main_markup(message))
    wb.close()
    fd = open('my_report.xlsx', 'rb')
    bot.send_document(message.chat.id, fd)
    fd.close()


def make_my_sheet1(message: types.Message, date_str: str, ws: Workbook.worksheets):
    current_date = datetime.datetime.today()
    start_date = datetime.datetime.strptime(date_str, "%d-%m-%Y")
    date = start_date
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 25
    i = 2
    j = 1
    user = users.get_user(message.chat.id)
    name = user[2] + " " + user[1]
    ws.cell(row=i, column=j, value=i - 1)
    ws.cell(row=i, column=j + 1, value=name)
    j = 3
    while date < current_date:
        row = visiting_fact.get_fact_by_date_and_id(user[0], date.strftime("%d-%m-%Y"))
        if not row:
            date += datetime.timedelta(days=1)
            continue
        i = 1
        ws.column_dimensions[get_column_letter(j)].width = 11
        ws.cell(row=i, column=j, value=date.strftime("%d-%m-%Y")).alignment = Alignment(horizontal="center",
                                                                                        vertical="center")
        i = 2
        if row[0] == 1:
            answer = "-"
            plan_row = visiting_fact.get_plan_answer_by_date_and_id(user[0], date.strftime("%d-%m-%Y"))
            if plan_row and plan_row[0] == 2:
                pf = PatternFill('solid', fgColor="F4CA16")
            else:
                pf = PatternFill('solid', fgColor="D7002F")
        else:
            answer = "+"
            pf = PatternFill('solid', fgColor="50C878")
        ws.cell(row=i, column=j, value=answer).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=i, column=j).fill = pf
        j += 1
        date += datetime.timedelta(days=1)

    max_column = ws.max_column
    max_row = ws.max_row
    ws.column_dimensions[get_column_letter(max_column+1)].width = 13
    ws.column_dimensions[get_column_letter(max_column+2)].width = 13
    ws.cell(row=1, column=max_column + 1, value="Посещений").alignment = Alignment(horizontal="center",
                                                                                     vertical="center")
    ws.cell(row=1, column=max_column + 2, value="Пропусков").alignment = Alignment(horizontal="center",
                                                                                     vertical="center")
    for i in range (2, max_row + 1):
        count_here = 0
        count_not_here = 0
        for j in range (3, max_column + 1):
            if ws.cell(row=i, column=j).value == "+":
                count_here += 1
            elif ws.cell(row=i, column=j).value == "-":
                count_not_here += 1
        ws.cell(row=i, column=max_column+1, value=count_here)
        ws.cell(row=i, column=max_column+2, value=count_not_here)

    i = max_row + 2
    s1 = "Зелёный фон: был на занятии"
    s2 = "Жёлтый фон: отсутствие с причиной"
    s3 = "Красный фон: отсутствие без причины"
    ws.cell(row=i, column=2, value=s1).font = Font(color="50C878")
    ws.cell(row=i+1, column=2, value=s2).font = Font(color="F4CA16")
    ws.cell(row=i+2, column=2, value=s3).font = Font(color="D7002F")

def make_my_sheet2(message: types.Message, date_str: str, ws: Workbook.worksheets):
    current_date = datetime.datetime.today()
    start_date = datetime.datetime.strptime(date_str, "%d-%m-%Y")
    date = start_date
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 25
    i = 2
    j = 1
    user = users.get_user(message.chat.id)
    name = user[2] + " " + user[1]
    ws.cell(row=i, column=j, value=i - 1)
    ws.cell(row=i, column=j + 1, value=name)
    j = 3
    while date < current_date:
        fact_row = visiting_fact.get_fact_by_date_and_id(user[0], date.strftime("%d-%m-%Y"))
        if not fact_row:
            date += datetime.timedelta(days=1)
            continue
        i = 1
        ws.column_dimensions[get_column_letter(j)].width = 15
        i = 2
        if fact_row[0] == 1:  # отсутствовал
            plan_row = visiting_fact.get_plan_answer_by_date_and_id(user[0], date.strftime("%d-%m-%Y"))
            pf = PatternFill('solid', fgColor="F4CA16")
            if plan_row and plan_row[0] == 2:
                skip_reason = plan_row[1]
            else:
                skip_reason = "-"
                pf = PatternFill('solid', fgColor="D7002F")
            ws.cell(row=i, column=j, value=skip_reason).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=i, column=j, value=skip_reason).fill = pf
            ws.cell(row=1, column=j, value=date.strftime("%d-%m-%Y")).alignment = Alignment(horizontal="center",
                                                                                            vertical="center")
        j += 1
        date += datetime.timedelta(days=1)


def valid_date(date_str: str) -> bool:
    try:
        datetime.datetime.strptime(date_str, '%d-%m-%Y')
        return True
    except ValueError:
        return False


def get_plan_report(message: types.Message):
    current_date = datetime.datetime.today().date()
    end_date = current_date + datetime.timedelta(days=7)
    date = current_date
    current_schedule = train_schedule.get_current_schedule()

    wb = Workbook()
    ws = wb.create_sheet("План посещений на неделю вперед", 0)
    wb.remove(wb['Sheet'])

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 25
    i = 2
    j = 1
    active_users = users.get_active_users_sorted_inBase()
    if not active_users:
        bot.send_message(message.chat.id, "В базе данных нет активных пользователей",
                         reply_markup=markups.main_markup(message))
        return
    for user in active_users:
        name = user[2] + " " + user[1]
        ws.cell(row=i, column=j, value=i - 1)
        ws.cell(row=i, column=j + 1, value=name)
        if user[9] == 1:
            ws.cell(row=i, column=j + 1).fill = PatternFill('solid', fgColor="63C5DA")
        i += 1

    j = 3
    while date <= end_date:
        weekday_number = date.weekday()+1
        status = False
        for day in current_schedule:
            if day[0] == weekday_number:
                status = True
                break
        if not status: #not a train day
            date += datetime.timedelta(days=1)
            continue
        i = 1
        ws.column_dimensions[get_column_letter(j)].width = 15
        ws.cell(row=i, column=j, value=date.strftime("%d-%m-%Y")).alignment = Alignment(horizontal="center",
                                                                                        vertical="center")
        i = 2
        for user in active_users:
            plan_row = visiting_fact.get_plan_answer_by_date_and_id(user[0], date.strftime("%d-%m-%Y"))
            if plan_row:
                if plan_row[0] == 2: #will not come
                    skip_reason = plan_row[1]
                    pf = PatternFill('solid', fgColor="F4CA16") 
                else: #will come
                    skip_reason = "+"
                    pf = PatternFill('solid', fgColor="50C878")
            else:
                skip_reason = "-"
                pf = PatternFill('solid', fgColor="D7002F")
            ws.cell(row=i, column=j, value=skip_reason).alignment = Alignment(horizontal="center",
                                                                              vertical="center")
            ws.cell(row=i, column=j, value=skip_reason).fill = pf
            i += 1
        j += 1
        date += datetime.timedelta(days=1)

    i = ws.max_row + 2
    s1 = "Зелёный фон: будет на занятии"
    s2 = "Жёлтый фон: отсутствие с причиной"
    s3 = "Красный фон: план не заполнен"
    s4 = "Голубой фон: в основном составе"
    ws.cell(row=i, column=2, value=s1).font = Font(color="50C878")
    ws.cell(row=i + 1, column=2, value=s2).font = Font(color="F4CA16")
    ws.cell(row=i + 2, column=2, value=s3).font = Font(color="D7002F")
    ws.cell(row=i + 3, column=2, value=s4).font = Font(color="63C5DA")

    wb.save("plan_report.xlsx")
    bot.send_message(message.chat.id, string_constants.GET_PLAN_REPORT + ":",
                     reply_markup=markups.main_markup(message))
    wb.close()
    fd = open('plan_report.xlsx', 'rb')
    bot.send_document(message.chat.id, fd)
    fd.close()

